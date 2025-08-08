import os
from tkinter import Tk, filedialog, Button, Label, Text, Scrollbar
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

pd.set_option('future.no_silent_downcasting', True)

list_of_non_materials = ['Материалы', 'Фурнитура на монтаж', 'Фурнитура']
list_of_furniture = ['петля', 'направляющие', 'планка', 'заглушка', 'комплект', 'замок']
list_of_LKM = ['ral', 'ncs', 'отвердитель', 'разбавитель', 'грунт', 'лак', 'эмаль', 'порошковая', 'краска']
list_of_mirror = ['стекло', 'зеркало']
list_of_still = ['лист', 'труба']
list_of_wood = ['лдсп', 'акрил', 'кромка', 'мдф', 'фанера', 'дсп', 'хдф']


# Функция для объединения файлов Excel
def merge_excel_files():
    global unit
    units = []

    # Выбор нескольких файлов
    file_paths = filedialog.askopenfilenames(
        title="Выберите файлы Excel",
        filetypes=(("Excel files", "*.xls*"), ("All Files", "*.*"))
    )

    if not file_paths:
        return

    try:
        df_list = []

        # Проход по каждому файлу
        for path in file_paths:
            df = pd.read_excel(path)

            # Поиск строки с "Артикул изделия"
            for idx, row in df.iterrows():
                values = row.to_list()
                if any(cell == 'Артикул изделия' for cell in row):
                    pos = values.index('Артикул изделия')
                    if len(values) > pos + 1:
                        if values[pos + 1]:
                            unit = values[pos + 1]
                        else:
                            unit = " "
                        if unit not in units:
                            units.append(unit)
                    else:
                        continue
                    break

            # Поиск строки с "Артикул"
            start_row_idx = None
            for idx, row in df.iterrows():
                if any(cell == 'Артикул' for cell in row):
                    start_row_idx = idx
                    break

            if start_row_idx is None:
                raise Exception(f"В файле '{path}' не найдено ключевое слово 'Артикул'.")

            # Зацепились за следующую строку после "Артикул" как шапку таблицы
            headers = df.iloc[start_row_idx].values.tolist()

            # Проверка наличия нужных колонок
            required_columns = ["Артикул", "Наименование материала", "Ед. изм.", "Количество в заказе", "Примечание"]
            missing_cols = set(required_columns) - set(headers)
            if missing_cols:
                raise Exception(f"В файле '{path}' не хватает колонок: {missing_cols}")

            # Читаем весь Excel-файл
            df = pd.read_excel(path)

            # Выделяем нужный диапазон данных
            start_row_idx += 1  # Первая строка данных
            useful_data = df.iloc[start_row_idx:, :].copy(deep=True)  # Полностью скопировали таблицу

            # Назначаем названия колонок

            useful_data.columns = headers

            # Фиксируем пустые значения в колонке "Артикул"
            useful_data["Артикул"] = useful_data["Артикул"].fillna(" ")
            useful_data["Примечание"] = useful_data["Примечание"].fillna(" ")

            # Заменяем пустые значения в колонке "Количество в заказе" на 0
            temp_col = useful_data["Количество в заказе"].fillna(0)
            result = temp_col.infer_objects(copy=False)
            useful_data["Количество в заказе"] = result

            # Заменяем единицы измерения если ячейка пустая
            useful_data["Ед. изм."] = useful_data["Ед. изм."].fillna("н/а")

            # Исключение строк со значениями из списка-фильтра
            useless_rows_mask = useful_data['Наименование материала'].isin(list_of_non_materials)
            useful_data = useful_data[~useless_rows_mask].reset_index(drop=True)

            # Оставляем только требуемые колонки
            useful_data = useful_data[required_columns]

            useful_data["Номер заказа"] = unit

            # Добавляем полученный DataFrame в список результатов
            df_list.append(useful_data)

        # Объединение всех собранных таблиц
        merged_df = pd.concat(df_list)

        merged_df['Номер заказа'] = merged_df['Номер заказа'].astype(str)
        merged_df['Артикул'] = merged_df['Артикул'].astype(str)

        grouped_df = merged_df.groupby('Наименование материала').agg({
            'Номер заказа': lambda x: ', '.join(x.unique()),
            'Артикул': lambda x: ', '.join(x.unique()),
            'Наименование материала': 'first',
            'Ед. изм.': 'first',
            'Количество в заказе': 'sum',
            'Примечание': lambda x: ', '.join(x.unique())
        })

        # Путь для сохранения итогового файла
        output_path = os.path.join(os.getcwd(), 'Сводный заказ.xlsx')

        info_df = pd.DataFrame({'Сводная таблица': [units]}, index=[0])

        # Преобразуем итоговый объединенный DataFrame в объекты Excel
        wb = Workbook()
        ws = wb.active

        if output_path:

            # Создаем новый dataframe для записи заказа
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Экспортируем информацию о заказе
                info_df.to_excel(writer, sheet_name='Итоги', index=False, startrow=0)

                worksheet = writer.sheets['Итоги']
                # Переносимся ниже и выводим итоговые материалы
                grouped_df.to_excel(writer, sheet_name='Итоги', index=False, startrow=len(info_df) + 2)

                for r in dataframe_to_rows(grouped_df, index=False, header=True):
                    ws.append(r)
                start_row = len(info_df) + 2
                red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                rows_with_data = list(worksheet.iter_rows(min_row=start_row + 1))
                for i, row in enumerate(rows_with_data):
                    values_in_row = [cell.value for cell in row]
                    if 'н/а' in map(str, values_in_row):
                        for cell in row:
                            cell.fill = red_fill
                wb.save(output_path)

            result_text.delete('1.0', 'end')
            result_text.insert('end', f'Файл успешно сохранён в {output_path}')
        else:
            raise Exception('Отмена сохранения файла.')

    except FileNotFoundError:
        result_text.delete('1.0', 'end')
        result_text.insert('end', 'Ошибка: Один или несколько файлов не найдены.')
    except KeyError:
        result_text.delete('1.0', 'end')
        result_text.insert('end', 'Ошибка: Неверные названия колонок в файлах.')
    except Exception as e:
        result_text.delete('1.0', 'end')
        result_text.insert('end', f'Ошибка: {e}')


def merge_merged_files():
    file_paths = filedialog.askopenfilenames(
        title="Выберите файлы Excel",
        filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
    )

    if not file_paths:
        result_text.delete('1.0', 'end')
        result_text.insert('end', f"Файлы не выбраны")
        return None

    # Чтение всех выбранных файлов и обработка данных
    df_list = []
    for path in file_paths:
        try:
            df = pd.read_excel(path, skiprows=3)  # Пропускаем первые три строки

            # Проверяем наличие нужных столбцов
            required_columns = ['Номер заказа', 'Артикул', 'Наименование материала',
                                'Ед. изм.', 'Количество в заказе', 'Примечание']
            missing_cols = set(required_columns).difference(df.columns)
            if len(missing_cols):
                raise ValueError(f"В файле {path} отсутствуют необходимые столбцы: {missing_cols}")

            # Преобразуем количество в числовой формат
            df['Количество в заказе'] = pd.to_numeric(df['Количество в заказе'], errors='coerce')

            # Удаляем дублирующиеся строки по артикулу и наименованию материала
            df.drop_duplicates(subset=['Артикул', 'Наименование материала'], inplace=True)

            df_list.append(df)
        except Exception as e:
            result_text.delete('1.0', 'end')
            result_text.insert('end', f"Произошла ошибка при обработке файла. {e}")

    # Объединение всех полученных таблиц
    merged_df = pd.concat(df_list, ignore_index=True)

    merged_df['Номер заказа'] = merged_df['Номер заказа'].astype(str)
    merged_df['Артикул'] = merged_df['Артикул'].fillna(" ")
    merged_df['Артикул'] = merged_df['Артикул'].astype(str)
    merged_df['Примечание'] = merged_df['Примечание'].fillna(" ")

    grouped_df = merged_df.groupby('Наименование материала').agg({
        'Номер заказа': lambda x: ', '.join(sorted(set(', '.join(x).split(', ')))),
        'Артикул': lambda x: ', '.join(x.unique()),
        'Наименование материала': 'first',
        'Ед. изм.': 'first',
        'Количество в заказе': 'sum',
        'Примечание': lambda x: ', '.join(x.unique())
    })

    # Создаем красную заливку для условного форматирования
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Цвет светло-красный

    # Сохраняем результат в новый Excel-файл
    output_path = os.path.join(os.getcwd(), 'Объединенный файл.xlsx')
    if not output_path:
        result_text.delete('1.0', 'end')
        result_text.insert('end', f"Не получилось сохранить файл")
        return None

    workbook = Workbook()
    sheet = workbook.active

    # Запись заголовков
    headers = list(grouped_df.columns)
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header

    # Добавляем таблицу в лист
    rows = dataframe_to_rows(grouped_df, index=False, header=False)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value

        # Получаем наименование материала и количество текущего ряда
        material_name = str(
            sheet.cell(row=row_idx, column=grouped_df.columns.get_loc('Наименование материала') + 1).value)
        quantity = float(sheet.cell(row=row_idx, column=grouped_df.columns.get_loc('Количество в заказе') + 1).value)
        unit_measure = str(sheet.cell(row=row_idx, column=grouped_df.columns.get_loc('Ед. изм.') + 1).value)

        # Проверка на красные строки
        if (
                material_name.strip() in list_of_non_materials or  # Название материала входит в чёрный список
                quantity <= 0 or  # Количество меньше или равно нулю
                unit_measure.lower().strip() == "н/а"  # Единица измерения отсутствует
        ):
            for cell in sheet[row_idx]:
                cell.fill = red_fill
    try:
        workbook.save(output_path)
        result_text.delete('1.0', 'end')
        result_text.insert('end', f"Создание объединённого файла успешно завершилось! Файл сохранён в '{output_path}'")
    except Exception as e:
        result_text.delete('1.0', 'end')
        result_text.insert('end', f"Ошибка доступа к файлу {e}")


root = Tk()
root.title("Объединение Excel-файлов")

select_button = Button(root, text="Выбрать файлы", command=merge_excel_files)
select_orders = Button(root, text="Выбрать заказы", command=merge_merged_files)
select_button.pack(pady=10)
select_orders.pack(pady=10)

result_label = Label(root, text="Результат:")
result_label.pack()

scrollbar = Scrollbar(root)
scrollbar.pack(side='right', fill='y')

result_text = Text(root, yscrollcommand=scrollbar.set, height=10, width=80)
result_text.pack(fill='both', expand=True)
scrollbar.config(command=result_text.yview)

root.mainloop()
